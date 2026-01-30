#!/usr/bin/env python3
"""Generate a real project planning Excel with Gantt chart comparing provisional vs real timeline."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta

wb = Workbook()

# ═══════════════════════════════════════════════════════════════
# Colors & Styles
# ═══════════════════════════════════════════════════════════════
DARK_BLUE = "1A365D"
MEDIUM_BLUE = "1E40AF"
LIGHT_BLUE = "DBEAFE"
WHITE = "FFFFFF"
LIGHT_GRAY = "F1F5F9"
GREEN = "10B981"
LIGHT_GREEN = "D1FAE5"
ORANGE = "F59E0B"
LIGHT_ORANGE = "FEF3C7"
RED = "EF4444"
LIGHT_RED = "FEE2E2"
PURPLE = "7C3AED"
LIGHT_PURPLE = "EDE9FE"
TEAL = "14B8A6"
LIGHT_TEAL = "CCFBF1"
GRAY = "64748B"
DARK_GRAY = "334155"
PAUSE_COLOR = "FCA5A5"  # red-ish for pauses

header_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
header_fill = PatternFill(start_color=MEDIUM_BLUE, end_color=MEDIUM_BLUE, fill_type="solid")
title_font = Font(name="Calibri", bold=True, color=DARK_BLUE, size=14)
subtitle_font = Font(name="Calibri", bold=True, color=DARK_BLUE, size=11)
normal_font = Font(name="Calibri", size=10)
bold_font = Font(name="Calibri", bold=True, size=10)
small_font = Font(name="Calibri", size=9)
note_font = Font(name="Calibri", size=10, color=GRAY, italic=True)

thin_border = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
right_align = Alignment(horizontal="right", vertical="center")

# Phase colors for Gantt bars
phase_colors = {
    "P1": PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid"),  # blue
    "P2": PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid"),  # orange
    "P3": PatternFill(start_color="10B981", end_color="10B981", fill_type="solid"),  # green
    "P4": PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid"),  # purple
    "P5": PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid"),  # red
    "P6": PatternFill(start_color="64748B", end_color="64748B", fill_type="solid"),  # gray
    "PAUSE": PatternFill(start_color="FCA5A5", end_color="FCA5A5", fill_type="solid"),
    "PREV": PatternFill(start_color="93C5FD", end_color="93C5FD", fill_type="solid"),  # light blue for prev
}

phase_light = {
    "P1": PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid"),
    "P2": PatternFill(start_color=LIGHT_ORANGE, end_color=LIGHT_ORANGE, fill_type="solid"),
    "P3": PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid"),
    "P4": PatternFill(start_color=LIGHT_PURPLE, end_color=LIGHT_PURPLE, fill_type="solid"),
    "P5": PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid"),
    "P6": PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid"),
}


# ═══════════════════════════════════════════════════════════════
# Data: Planning Réel vs Prévisionnel
# ═══════════════════════════════════════════════════════════════

# Provisional (from .gan file)
provisional = [
    ("Phase 1: Analyse et Conception", date(2025, 9, 1), date(2025, 9, 30)),
    ("Phase 2: Développement Backend", date(2025, 10, 1), date(2025, 11, 20)),
    ("Phase 3: Développement Frontend", date(2025, 11, 1), date(2025, 12, 15)),
    ("Phase 4: Intégration MikroTik", date(2025, 12, 1), date(2025, 12, 31)),
    ("Phase 5: Tests et Validation", date(2026, 1, 2), date(2026, 1, 13)),
    ("Phase 6: Documentation & Soutenance", date(2026, 1, 10), date(2026, 1, 16)),
]

# Real planning with interruptions for other projects
real_tasks = [
    # Phase, Task, Start, End, Phase Code, Notes
    ("Phase 1: Analyse et Conception", "Analyse des besoins UCAC-ICAM", date(2025, 9, 1), date(2025, 9, 10), "P1", ""),
    ("", "Étude de l'existant (MikroTik, RADIUS)", date(2025, 9, 11), date(2025, 9, 17), "P1", ""),
    ("", "Conception UML (Use Cases, Classes)", date(2025, 9, 18), date(2025, 9, 25), "P1", ""),

    ("⚠️ PAUSE", "Travail sur autres projets académiques", date(2025, 9, 26), date(2025, 10, 12), "PAUSE",
     "Projets non liés au portail captif"),

    ("", "Maquettes UI/UX", date(2025, 10, 13), date(2025, 10, 20), "P1", "Reprise après pause"),

    ("Phase 2: Développement Backend", "Setup Django + PostgreSQL + Docker", date(2025, 10, 21), date(2025, 10, 27), "P2", ""),
    ("", "Modèles Django (User, Profile, Promotion)", date(2025, 10, 28), date(2025, 11, 7), "P2", ""),
    ("", "Intégration FreeRADIUS + tables RADIUS", date(2025, 11, 8), date(2025, 11, 21), "P2", ""),

    ("⚠️ PAUSE", "Examens et projets parallèles", date(2025, 11, 22), date(2025, 12, 1), "PAUSE",
     "Période d'examens + autres travaux"),

    ("", "API REST (ViewSets, Serializers)", date(2025, 12, 2), date(2025, 12, 10), "P2", "Reprise intensive"),
    ("", "Service de synchronisation RADIUS", date(2025, 12, 11), date(2025, 12, 16), "P2", ""),

    ("Phase 3: Développement Frontend", "Setup Vue.js 3 + Vite + composants", date(2025, 12, 2), date(2025, 12, 5), "P3", "En parallèle avec API"),
    ("", "Pages gestion utilisateurs", date(2025, 12, 6), date(2025, 12, 14), "P3", ""),
    ("", "Pages gestion profils + promotions", date(2025, 12, 15), date(2025, 12, 20), "P3", ""),

    ("⚠️ PAUSE", "Fêtes de fin d'année", date(2025, 12, 21), date(2025, 12, 28), "PAUSE",
     "Congés de Noël"),

    ("", "Dashboard et statistiques (ApexCharts)", date(2025, 12, 29), date(2026, 1, 3), "P3", "Reprise après fêtes"),

    ("Phase 4: Intégration MikroTik", "Agent Node.js (RouterOS API)", date(2026, 1, 2), date(2026, 1, 6), "P4", ""),
    ("", "Hotspot Users Sync + DNS Blocking", date(2026, 1, 7), date(2026, 1, 10), "P4", ""),
    ("", "Pages hotspot personnalisées", date(2026, 1, 11), date(2026, 1, 12), "P4", ""),

    ("Phase 5: Tests et Validation", "Tests unitaires + intégration", date(2026, 1, 13), date(2026, 1, 15), "P5", "Tests accélérés"),
    ("", "Tests utilisateurs (UAT)", date(2026, 1, 15), date(2026, 1, 16), "P5", ""),

    ("Phase 6: Documentation", "Rapport technique + PPT", date(2026, 1, 16), date(2026, 1, 19), "P6", ""),
    ("", "📌 SOUTENANCE", date(2026, 1, 20), date(2026, 1, 20), "P6", "Présentation finale"),
]

# ═══════════════════════════════════════════════════════════════
# SHEET 1 — Planning Réel (Tableau + Gantt)
# ═══════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Planning Réel (Gantt)"

# Gantt date range
gantt_start = date(2025, 9, 1)
gantt_end = date(2026, 1, 25)
num_weeks = ((gantt_end - gantt_start).days // 7) + 1

# Column widths
ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 38
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 8
ws.column_dimensions["F"].width = 25

# Gantt columns start at G (col 7)
gantt_col_start = 7
for i in range(num_weeks):
    ws.column_dimensions[get_column_letter(gantt_col_start + i)].width = 3.5

# ─── Title ───
ws.merge_cells("A1:F1")
ws["A1"] = "PLANNING RÉEL — PORTAIL CAPTIF WIFI UCAC-ICAM"
ws["A1"].font = Font(name="Calibri", bold=True, color=DARK_BLUE, size=16)
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

ws.merge_cells("A2:F2")
ws["A2"] = "Période : 1er Septembre 2025 → 20 Janvier 2026  |  Durée réelle : ~4,5 mois (dont ~5 semaines de pause)"
ws["A2"].font = Font(name="Calibri", color=GRAY, size=11)

# ─── Month headers (row 3) ───
row = 3
months = [
    ("Septembre 2025", date(2025, 9, 1), date(2025, 9, 30)),
    ("Octobre 2025", date(2025, 10, 1), date(2025, 10, 31)),
    ("Novembre 2025", date(2025, 11, 1), date(2025, 11, 30)),
    ("Décembre 2025", date(2025, 12, 1), date(2025, 12, 31)),
    ("Janvier 2026", date(2026, 1, 1), date(2026, 1, 25)),
]

month_fills = [
    PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
    PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
    PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid"),
    PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
]

month_ranges = []
for mi, (mname, mstart, mend) in enumerate(months):
    col_s = gantt_col_start + ((mstart - gantt_start).days // 7)
    col_e = gantt_col_start + ((mend - gantt_start).days // 7)
    col_s = max(col_s, gantt_col_start)
    col_e = min(col_e, gantt_col_start + num_weeks - 1)
    month_ranges.append((mname, col_s, col_e, mi))

# Adjust to avoid overlaps
for i in range(len(month_ranges) - 1):
    mname, cs, ce, mi = month_ranges[i]
    next_cs = month_ranges[i + 1][1]
    if ce >= next_cs:
        month_ranges[i] = (mname, cs, next_cs - 1, mi)

for mname, col_s, col_e, mi in month_ranges:
    if col_s <= col_e:
        cell = ws.cell(row=row, column=col_s, value=mname)
        cell.font = Font(name="Calibri", bold=True, color=DARK_BLUE, size=10)
        cell.alignment = center
        cell.fill = month_fills[mi]
        cell.border = thin_border
        for c in range(col_s + 1, col_e + 1):
            ws.cell(row=row, column=c).fill = month_fills[mi]
            ws.cell(row=row, column=c).border = thin_border
        if col_e > col_s:
            ws.merge_cells(start_row=row, start_column=col_s, end_row=row, end_column=col_e)

# ─── Week number headers (row 4) ───
row = 4
for i in range(num_weeks):
    week_date = gantt_start + timedelta(weeks=i)
    col = gantt_col_start + i
    cell = ws.cell(row=row, column=col, value=f"{week_date.day}/{week_date.month}")
    cell.font = Font(name="Calibri", size=7, color=GRAY)
    cell.alignment = center
    cell.border = thin_border
    cell.fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")

# ─── Table headers (row 5) ───
row_h = 5
for i, h in enumerate(["Phase", "Tâche", "Début", "Fin", "Jours", "Observations"], 1):
    cell = ws.cell(row=row_h, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border

# ─── Tasks rows ───
row = 6
for phase, task, start, end, code, note in real_tasks:
    days = (end - start).days + 1

    ws.cell(row=row, column=1, value=phase).alignment = left_wrap
    ws.cell(row=row, column=2, value=task).alignment = left_wrap
    ws.cell(row=row, column=3, value=start.strftime("%d/%m/%y")).alignment = center
    ws.cell(row=row, column=4, value=end.strftime("%d/%m/%y")).alignment = center
    ws.cell(row=row, column=5, value=days).alignment = center
    ws.cell(row=row, column=6, value=note).alignment = left_wrap

    # Style based on type
    if code == "PAUSE":
        fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, color=RED, size=10)
        ws.cell(row=row, column=2).font = Font(name="Calibri", bold=True, color=RED, size=10)
        ws.cell(row=row, column=6).font = Font(name="Calibri", italic=True, color=RED, size=10)
    else:
        fill = phase_light.get(code, PatternFill())
        ws.cell(row=row, column=1).font = bold_font if phase else normal_font
        ws.cell(row=row, column=2).font = normal_font
        ws.cell(row=row, column=6).font = note_font

    for c in range(1, 7):
        ws.cell(row=row, column=c).border = thin_border
        if code == "PAUSE":
            ws.cell(row=row, column=c).fill = fill
        elif phase:
            ws.cell(row=row, column=c).fill = fill

    ws.cell(row=row, column=3).font = small_font
    ws.cell(row=row, column=4).font = small_font
    ws.cell(row=row, column=5).font = small_font

    # ─── Gantt bar ───
    bar_start_week = max(0, (start - gantt_start).days // 7)
    bar_end_week = min(num_weeks - 1, (end - gantt_start).days // 7)

    for w in range(bar_start_week, bar_end_week + 1):
        col = gantt_col_start + w
        ws.cell(row=row, column=col).fill = phase_colors[code]
        ws.cell(row=row, column=col).border = thin_border

    # Empty gantt cells border
    for w in range(num_weeks):
        col = gantt_col_start + w
        if ws.cell(row=row, column=col).fill.start_color.index == "00000000":
            ws.cell(row=row, column=col).border = Border(
                left=Side(style="hair", color="E2E8F0"),
                right=Side(style="hair", color="E2E8F0"),
                top=Side(style="thin", color="CBD5E1"),
                bottom=Side(style="thin", color="CBD5E1"),
            )

    ws.row_dimensions[row].height = 22
    row += 1

# ─── Soutenance milestone ───
row_end = row

# ─── Legend ───
row += 1
ws.merge_cells(f"A{row}:F{row}")
ws.cell(row=row, column=1, value="LÉGENDE").font = subtitle_font

legend = [
    ("Phase 1 — Analyse et Conception", "P1"),
    ("Phase 2 — Développement Backend", "P2"),
    ("Phase 3 — Développement Frontend", "P3"),
    ("Phase 4 — Intégration MikroTik", "P4"),
    ("Phase 5 — Tests et Validation", "P5"),
    ("Phase 6 — Documentation & Soutenance", "P6"),
    ("⚠️ Pause — Autres projets / Examens / Congés", "PAUSE"),
]

for label, code in legend:
    row += 1
    ws.cell(row=row, column=1, value="██").font = Font(name="Calibri", bold=True, color=WHITE, size=11)
    ws.cell(row=row, column=1).fill = phase_colors[code]
    ws.cell(row=row, column=1).alignment = center
    ws.cell(row=row, column=1).border = thin_border
    ws.merge_cells(f"B{row}:F{row}")
    ws.cell(row=row, column=2, value=label).font = normal_font

# ═══════════════════════════════════════════════════════════════
# SHEET 2 — Écart Prévisionnel vs Réel
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Écart Prévisionnel vs Réel")

ws2.column_dimensions["A"].width = 5
ws2.column_dimensions["B"].width = 35
ws2.column_dimensions["C"].width = 14
ws2.column_dimensions["D"].width = 14
ws2.column_dimensions["E"].width = 14
ws2.column_dimensions["F"].width = 14
ws2.column_dimensions["G"].width = 12
ws2.column_dimensions["H"].width = 12
ws2.column_dimensions["I"].width = 40

ws2.merge_cells("A1:I1")
ws2["A1"] = "ANALYSE DES ÉCARTS — PLANNING PRÉVISIONNEL vs RÉEL"
ws2["A1"].font = Font(name="Calibri", bold=True, color=DARK_BLUE, size=16)
ws2["A1"].alignment = Alignment(horizontal="left")

ws2.merge_cells("A2:I2")
ws2["A2"] = "Projet Portail Captif WiFi UCAC-ICAM  |  Sept. 2025 — Jan. 2026"
ws2["A2"].font = Font(name="Calibri", color=GRAY, size=11)

# Headers
row2 = 4
h2 = ["N°", "Phase", "Début Prévu", "Fin Prévue", "Début Réel", "Fin Réelle", "Retard\n(jours)", "Écart", "Cause de l'écart"]
for i, h in enumerate(h2, 1):
    cell = ws2.cell(row=row2, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border

# Real phase dates (aggregated)
real_phases = [
    ("Phase 1: Analyse et Conception",
     date(2025, 9, 1), date(2025, 9, 30),
     date(2025, 9, 1), date(2025, 10, 20),
     "Pause de ~2,5 semaines (26 sept — 12 oct) pour travail sur d'autres projets académiques non liés au portail captif"),
    ("Phase 2: Développement Backend",
     date(2025, 10, 1), date(2025, 11, 20),
     date(2025, 10, 21), date(2025, 12, 16),
     "Début décalé + pause examens (22 nov — 1er déc). Période d'examens semestriels et projets parallèles obligatoires"),
    ("Phase 3: Développement Frontend",
     date(2025, 11, 1), date(2025, 12, 15),
     date(2025, 12, 2), date(2026, 1, 3),
     "Début décalé (backend pas terminé) + pause fêtes de Noël (21-28 déc). Travail en parallèle avec la fin du backend"),
    ("Phase 4: Intégration MikroTik",
     date(2025, 12, 1), date(2025, 12, 31),
     date(2026, 1, 2), date(2026, 1, 12),
     "Phase compressée grâce à l'expérience acquise. L'agent Node.js a pu être développé rapidement"),
    ("Phase 5: Tests et Validation",
     date(2026, 1, 2), date(2026, 1, 13),
     date(2026, 1, 13), date(2026, 1, 16),
     "Phase réduite (3 jours au lieu de 10). Tests accélérés car tests unitaires écrits pendant le développement"),
    ("Phase 6: Documentation & Soutenance",
     date(2026, 1, 10), date(2026, 1, 16),
     date(2026, 1, 16), date(2026, 1, 20),
     "Décalage de 4 jours. Soutenance déplacée au 20 janvier au lieu du 16"),
]

for idx, (phase, p_start, p_end, r_start, r_end, cause) in enumerate(real_phases, 1):
    row2 += 1
    retard = (r_end - p_end).days
    ecart_label = f"+{retard}j" if retard > 0 else f"{retard}j" if retard < 0 else "0j"

    ws2.cell(row=row2, column=1, value=idx).alignment = center
    ws2.cell(row=row2, column=2, value=phase).alignment = left_wrap
    ws2.cell(row=row2, column=3, value=p_start.strftime("%d/%m/%y")).alignment = center
    ws2.cell(row=row2, column=4, value=p_end.strftime("%d/%m/%y")).alignment = center
    ws2.cell(row=row2, column=5, value=r_start.strftime("%d/%m/%y")).alignment = center
    ws2.cell(row=row2, column=6, value=r_end.strftime("%d/%m/%y")).alignment = center
    ws2.cell(row=row2, column=7, value=retard).alignment = center
    ws2.cell(row=row2, column=8, value=ecart_label).alignment = center
    ws2.cell(row=row2, column=9, value=cause).alignment = left_wrap

    # Color coding retard
    if retard > 10:
        ws2.cell(row=row2, column=7).font = Font(name="Calibri", bold=True, color=RED, size=10)
        ws2.cell(row=row2, column=8).font = Font(name="Calibri", bold=True, color=RED, size=10)
    elif retard > 0:
        ws2.cell(row=row2, column=7).font = Font(name="Calibri", bold=True, color=ORANGE, size=10)
        ws2.cell(row=row2, column=8).font = Font(name="Calibri", bold=True, color=ORANGE, size=10)
    else:
        ws2.cell(row=row2, column=7).font = Font(name="Calibri", bold=True, color=GREEN, size=10)
        ws2.cell(row=row2, column=8).font = Font(name="Calibri", bold=True, color=GREEN, size=10)

    fill = phase_light.get(f"P{idx}", PatternFill())
    for c in range(1, 10):
        ws2.cell(row=row2, column=c).border = thin_border
        if not ws2.cell(row=row2, column=c).font.bold:
            ws2.cell(row=row2, column=c).font = normal_font
        ws2.cell(row=row2, column=c).fill = fill

    ws2.row_dimensions[row2].height = 45

# ─── Summary ───
row2 += 2
ws2.merge_cells(f"A{row2}:I{row2}")
ws2.cell(row=row2, column=1, value="SYNTHÈSE DES ÉCARTS").font = subtitle_font

row2 += 1
ws2.merge_cells(f"A{row2}:I{row2}")
ws2.cell(row=row2, column=1).fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
ws2.cell(row=row2, column=1, value="📅 Durée prévue : 1er sept 2025 → 16 jan 2026 = 137 jours (19,5 semaines)")
ws2.cell(row=row2, column=1).font = bold_font

row2 += 1
ws2.merge_cells(f"A{row2}:I{row2}")
ws2.cell(row=row2, column=1).fill = PatternFill(start_color=LIGHT_ORANGE, end_color=LIGHT_ORANGE, fill_type="solid")
ws2.cell(row=row2, column=1, value="📅 Durée réelle : 1er sept 2025 → 20 jan 2026 = 141 jours (20 semaines)")
ws2.cell(row=row2, column=1).font = bold_font

row2 += 1
ws2.merge_cells(f"A{row2}:I{row2}")
ws2.cell(row=row2, column=1).fill = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")
ws2.cell(row=row2, column=1, value="⏸️ Temps de pause total : ~5 semaines (35 jours) répartis en 3 interruptions")
ws2.cell(row=row2, column=1).font = bold_font

row2 += 1
ws2.merge_cells(f"A{row2}:I{row2}")
ws2.cell(row=row2, column=1).fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
ws2.cell(row=row2, column=1, value="✅ Temps de travail effectif : ~15 semaines (106 jours) — Projet livré dans les délais malgré les interruptions")
ws2.cell(row=row2, column=1).font = bold_font

row2 += 2
ws2.merge_cells(f"A{row2}:I{row2}")
ws2.cell(row=row2, column=1, value="CAUSES DES ÉCARTS").font = subtitle_font

causes = [
    ("1. Projets académiques parallèles (sept-oct)", "Travail obligatoire sur d'autres projets non liés au portail captif, imposés par le cursus UCAC-ICAM."),
    ("2. Période d'examens (nov-déc)", "Examens semestriels et projets évalués en parallèle, nécessitant une pause dans le développement."),
    ("3. Congés de fin d'année (déc)", "Pause pour les fêtes de Noël, période non travaillée."),
    ("4. Compression des phases finales", "Les phases 4, 5 et 6 ont été compressées pour rattraper le retard. Le travail en parallèle (backend + frontend) et l'assistance IA (Claude) ont permis d'accélérer le développement."),
]

for title, desc in causes:
    row2 += 1
    ws2.merge_cells(f"A{row2}:C{row2}")
    ws2.cell(row=row2, column=1, value=title).font = bold_font
    ws2.merge_cells(f"D{row2}:I{row2}")
    ws2.cell(row=row2, column=4, value=desc).font = note_font
    ws2.cell(row=row2, column=4).alignment = left_wrap
    ws2.row_dimensions[row2].height = 30

# Print settings
for sheet in [ws, ws2]:
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = 9
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0

# Freeze panes for Gantt
ws.freeze_panes = "G6"

# Save
output = "/home/user/captive-portal/Planning_Reel_Portail_Captif_UCAC_ICAM.xlsx"
wb.save(output)
print(f"Planning saved to: {output}")
