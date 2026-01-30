#!/usr/bin/env python3
"""Generate a professional budget Excel file for the Captive Portal project."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# ═══════════════════════════════════════════════════════════════
# Colors
# ═══════════════════════════════════════════════════════════════
DARK_BLUE = "1A365D"
MEDIUM_BLUE = "1E40AF"
LIGHT_BLUE = "DBEAFE"
WHITE = "FFFFFF"
LIGHT_GRAY = "F1F5F9"
GREEN_BG = "D1FAE5"
ORANGE_BG = "FEF3C7"
RED_BG = "FEE2E2"
PURPLE_BG = "EDE9FE"

header_font = Font(name="Calibri", bold=True, color=WHITE, size=12)
header_fill = PatternFill(start_color=MEDIUM_BLUE, end_color=MEDIUM_BLUE, fill_type="solid")
title_font = Font(name="Calibri", bold=True, color=DARK_BLUE, size=14)
subtitle_font = Font(name="Calibri", bold=True, color=DARK_BLUE, size=11)
normal_font = Font(name="Calibri", size=11)
bold_font = Font(name="Calibri", bold=True, size=11)
total_font = Font(name="Calibri", bold=True, color=WHITE, size=12)
total_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")

cat_fills = {
    "RH": PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid"),
    "HW": PatternFill(start_color=ORANGE_BG, end_color=ORANGE_BG, fill_type="solid"),
    "SW": PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid"),
    "FONC": PatternFill(start_color=PURPLE_BG, end_color=PURPLE_BG, fill_type="solid"),
    "gray": PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid"),
}

thin_border = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
right_align = Alignment(horizontal="right", vertical="center")

XAF_FORMAT = '#,##0" FCFA"'
PERCENT_FORMAT = '0.0%'
EUR_RATE = 656  # 1 EUR = 656 XAF


def style_range(ws, row, col_start, col_end, font=None, fill=None, alignment=None, border=None, number_format=None):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border
        if number_format:
            cell.number_format = number_format


# ═══════════════════════════════════════════════════════════════
# SHEET 1 — Budget Détaillé
# ═══════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Budget Détaillé"

# Column widths
ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 40
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 18
ws.column_dimensions["E"].width = 18
ws.column_dimensions["F"].width = 22
ws.column_dimensions["G"].width = 15

# Title
ws.merge_cells("A1:G1")
ws["A1"] = "BUDGET PRÉVISIONNEL — PORTAIL CAPTIF WIFI UCAC-ICAM"
ws["A1"].font = Font(name="Calibri", bold=True, color=DARK_BLUE, size=16)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("A2:G2")
ws["A2"] = "Projet de fin d'études — Janvier 2026  |  Devise : FCFA (1 EUR = 656 FCFA)"
ws["A2"].font = Font(name="Calibri", color="64748B", size=11)
ws["A2"].alignment = Alignment(horizontal="center")

# Headers row 4
row = 4
headers = ["N°", "Désignation", "Qté", "Prix Unitaire (FCFA)", "Montant (FCFA)", "Sous-Total", "%"]
for i, h in enumerate(headers, 1):
    cell = ws.cell(row=row, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border

# ─── RESSOURCES HUMAINES ───
row = 5
ws.merge_cells(f"A{row}:G{row}")
ws.cell(row=row, column=1, value="1. RESSOURCES HUMAINES")
style_range(ws, row, 1, 7, font=subtitle_font, fill=cat_fills["RH"], alignment=left_wrap, border=thin_border)

rh_items = [
    ("Chef de projet / Développeur principal", 5, 250000),
    ("Ingénieur réseau (consultant MikroTik)", 2, 200000),
    ("Designer UI/UX (interfaces portail)", 1, 150000),
    ("Testeur QA / Assurance qualité", 1, 100000),
    ("Rédacteur technique (documentation)", 1, 50000),
]

rh_start = row + 1
for idx, (name, qty, price) in enumerate(rh_items, 1):
    row += 1
    ws.cell(row=row, column=1, value=idx).alignment = center
    ws.cell(row=row, column=2, value=name).alignment = left_wrap
    ws.cell(row=row, column=3, value=qty).alignment = center
    ws.cell(row=row, column=4, value=price).number_format = XAF_FORMAT
    ws.cell(row=row, column=4).alignment = right_align
    ws.cell(row=row, column=5, value=qty * price).number_format = XAF_FORMAT
    ws.cell(row=row, column=5).alignment = right_align
    for c in range(1, 8):
        ws.cell(row=row, column=c).font = normal_font
        ws.cell(row=row, column=c).border = thin_border
        if idx % 2 == 0:
            ws.cell(row=row, column=c).fill = cat_fills["gray"]
rh_end = row

# Sous-total RH
row += 1
ws.cell(row=row, column=2, value="Sous-total Ressources Humaines").font = bold_font
ws.cell(row=row, column=5).alignment = right_align
rh_total = sum(q * p for _, q, p in rh_items)
ws.cell(row=row, column=6, value=rh_total).number_format = XAF_FORMAT
ws.cell(row=row, column=6).font = bold_font
ws.cell(row=row, column=6).alignment = right_align
style_range(ws, row, 1, 7, fill=cat_fills["RH"], border=thin_border)

# ─── MATÉRIEL / HARDWARE ───
row += 1
ws.merge_cells(f"A{row}:G{row}")
ws.cell(row=row, column=1, value="2. MATÉRIEL (HARDWARE)")
style_range(ws, row, 1, 7, font=subtitle_font, fill=cat_fills["HW"], alignment=left_wrap, border=thin_border)

hw_items = [
    ("Routeur MikroTik RB951Ui-2HnD", 1, 75000),
    ("Ordinateur portable (développement)", 1, 400000),
    ("Serveur Ubuntu (test/production)", 1, 150000),
    ("Points d'accès WiFi TP-Link", 2, 35000),
    ("Câbles réseau Cat6 (3m)", 10, 2500),
    ("Onduleur / UPS", 1, 85000),
    ("Switch réseau 8 ports", 1, 25000),
]

for idx, (name, qty, price) in enumerate(hw_items, 1):
    row += 1
    ws.cell(row=row, column=1, value=idx).alignment = center
    ws.cell(row=row, column=2, value=name).alignment = left_wrap
    ws.cell(row=row, column=3, value=qty).alignment = center
    ws.cell(row=row, column=4, value=price).number_format = XAF_FORMAT
    ws.cell(row=row, column=4).alignment = right_align
    ws.cell(row=row, column=5, value=qty * price).number_format = XAF_FORMAT
    ws.cell(row=row, column=5).alignment = right_align
    for c in range(1, 8):
        ws.cell(row=row, column=c).font = normal_font
        ws.cell(row=row, column=c).border = thin_border
        if idx % 2 == 0:
            ws.cell(row=row, column=c).fill = cat_fills["gray"]

row += 1
ws.cell(row=row, column=2, value="Sous-total Matériel").font = bold_font
hw_total = sum(q * p for _, q, p in hw_items)
ws.cell(row=row, column=6, value=hw_total).number_format = XAF_FORMAT
ws.cell(row=row, column=6).font = bold_font
ws.cell(row=row, column=6).alignment = right_align
style_range(ws, row, 1, 7, fill=cat_fills["HW"], border=thin_border)

# ─── LOGICIELS ───
row += 1
ws.merge_cells(f"A{row}:G{row}")
ws.cell(row=row, column=1, value="3. LOGICIELS")
style_range(ws, row, 1, 7, font=subtitle_font, fill=cat_fills["SW"], alignment=left_wrap, border=thin_border)

sw_items = [
    ("Ubuntu Server 22.04 LTS", 1, 0, "Open Source"),
    ("PostgreSQL 15", 1, 0, "Open Source"),
    ("FreeRADIUS 3.0", 1, 0, "Open Source"),
    ("Django 5.x + DRF", 1, 0, "Open Source"),
    ("Vue.js 3 + Vite", 1, 0, "Open Source"),
    ("Node.js (Agent MikroTik)", 1, 0, "Open Source"),
    ("Docker + Docker Compose", 1, 0, "Open Source"),
    ("Prometheus + Grafana", 1, 0, "Open Source"),
    ("Nom de domaine (.cm)", 1, 15000, "Annuel"),
    ("Certificat SSL (Let's Encrypt)", 1, 0, "Gratuit"),
]

for idx, (name, qty, price, note) in enumerate(sw_items, 1):
    row += 1
    ws.cell(row=row, column=1, value=idx).alignment = center
    ws.cell(row=row, column=2, value=name).alignment = left_wrap
    ws.cell(row=row, column=3, value=qty).alignment = center
    ws.cell(row=row, column=4, value=price).number_format = XAF_FORMAT
    ws.cell(row=row, column=4).alignment = right_align
    ws.cell(row=row, column=5, value=qty * price).number_format = XAF_FORMAT
    ws.cell(row=row, column=5).alignment = right_align
    for c in range(1, 8):
        ws.cell(row=row, column=c).font = normal_font
        ws.cell(row=row, column=c).border = thin_border
        if price == 0:
            ws.cell(row=row, column=5).value = "GRATUIT"
            ws.cell(row=row, column=5).number_format = "@"
            ws.cell(row=row, column=5).font = Font(name="Calibri", color="10B981", size=11, bold=True)
            ws.cell(row=row, column=5).alignment = center
            ws.cell(row=row, column=4).value = "GRATUIT"
            ws.cell(row=row, column=4).number_format = "@"
            ws.cell(row=row, column=4).font = Font(name="Calibri", color="10B981", size=11, bold=True)
            ws.cell(row=row, column=4).alignment = center
        if idx % 2 == 0:
            ws.cell(row=row, column=c).fill = cat_fills["gray"]

row += 1
ws.cell(row=row, column=2, value="Sous-total Logiciels").font = bold_font
sw_total = sum(q * p for _, q, p, _ in sw_items)
ws.cell(row=row, column=6, value=sw_total).number_format = XAF_FORMAT
ws.cell(row=row, column=6).font = bold_font
ws.cell(row=row, column=6).alignment = right_align
style_range(ws, row, 1, 7, fill=cat_fills["SW"], border=thin_border)

# ─── FONCTIONNEMENT ───
row += 1
ws.merge_cells(f"A{row}:G{row}")
ws.cell(row=row, column=1, value="4. FRAIS DE FONCTIONNEMENT")
style_range(ws, row, 1, 7, font=subtitle_font, fill=cat_fills["FONC"], alignment=left_wrap, border=thin_border)

fonc_items = [
    ("Connexion Internet (test)", 3, 15000),
    ("Électricité (serveur)", 6, 10000),
    ("Sauvegarde cloud", 12, 5000),
    ("Impressions / Documentation", 1, 20000),
    ("Déplacements / Transport", 5, 10000),
]

for idx, (name, qty, price) in enumerate(fonc_items, 1):
    row += 1
    ws.cell(row=row, column=1, value=idx).alignment = center
    ws.cell(row=row, column=2, value=name).alignment = left_wrap
    ws.cell(row=row, column=3, value=qty).alignment = center
    ws.cell(row=row, column=4, value=price).number_format = XAF_FORMAT
    ws.cell(row=row, column=4).alignment = right_align
    ws.cell(row=row, column=5, value=qty * price).number_format = XAF_FORMAT
    ws.cell(row=row, column=5).alignment = right_align
    for c in range(1, 8):
        ws.cell(row=row, column=c).font = normal_font
        ws.cell(row=row, column=c).border = thin_border
        if idx % 2 == 0:
            ws.cell(row=row, column=c).fill = cat_fills["gray"]

row += 1
ws.cell(row=row, column=2, value="Sous-total Fonctionnement").font = bold_font
fonc_total = sum(q * p for _, q, p in fonc_items)
ws.cell(row=row, column=6, value=fonc_total).number_format = XAF_FORMAT
ws.cell(row=row, column=6).font = bold_font
ws.cell(row=row, column=6).alignment = right_align
style_range(ws, row, 1, 7, fill=cat_fills["FONC"], border=thin_border)

# ─── TOTAL GÉNÉRAL ───
row += 2
total = rh_total + hw_total + sw_total + fonc_total

ws.merge_cells(f"A{row}:E{row}")
ws.cell(row=row, column=1, value="TOTAL GÉNÉRAL")
ws.cell(row=row, column=6, value=total).number_format = XAF_FORMAT
style_range(ws, row, 1, 7, font=total_font, fill=total_fill, alignment=center, border=thin_border)
ws.cell(row=row, column=6).alignment = right_align

# Contingency
row += 1
contingency = int(total * 0.1)
ws.merge_cells(f"A{row}:E{row}")
ws.cell(row=row, column=1, value="Provision pour imprévus (10%)")
ws.cell(row=row, column=1).font = bold_font
ws.cell(row=row, column=6, value=contingency).number_format = XAF_FORMAT
ws.cell(row=row, column=6).font = bold_font
ws.cell(row=row, column=6).alignment = right_align
style_range(ws, row, 1, 7, fill=cat_fills["gray"], border=thin_border)

# Grand total
row += 1
grand_total = total + contingency
ws.merge_cells(f"A{row}:E{row}")
ws.cell(row=row, column=1, value="BUDGET TOTAL (avec imprévus)")
ws.cell(row=row, column=6, value=grand_total).number_format = XAF_FORMAT
style_range(ws, row, 1, 7, font=total_font, fill=PatternFill(start_color="10B981", end_color="10B981", fill_type="solid"), alignment=center, border=thin_border)
ws.cell(row=row, column=6).alignment = right_align

# EUR equivalent
row += 1
ws.merge_cells(f"A{row}:E{row}")
ws.cell(row=row, column=1, value=f"Équivalent en EUR (1 EUR = {EUR_RATE} FCFA)")
ws.cell(row=row, column=1).font = Font(name="Calibri", italic=True, color="64748B", size=11)
ws.cell(row=row, column=6, value=round(grand_total / EUR_RATE, 2))
ws.cell(row=row, column=6).number_format = '#,##0.00" EUR"'
ws.cell(row=row, column=6).font = Font(name="Calibri", italic=True, color="64748B", size=11)
ws.cell(row=row, column=6).alignment = right_align

# ─── Percentages in column G ───
# Go back and fill percentages for subtotals
for r in range(5, row + 1):
    val = ws.cell(row=r, column=6).value
    if isinstance(val, (int, float)) and val > 0 and val != grand_total and val != contingency and val != round(grand_total / EUR_RATE, 2):
        ws.cell(row=r, column=7, value=val / total)
        ws.cell(row=r, column=7).number_format = PERCENT_FORMAT
        ws.cell(row=r, column=7).alignment = center
        ws.cell(row=r, column=7).font = bold_font

# ═══════════════════════════════════════════════════════════════
# SHEET 2 — Résumé
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Résumé")

ws2.column_dimensions["A"].width = 5
ws2.column_dimensions["B"].width = 35
ws2.column_dimensions["C"].width = 22
ws2.column_dimensions["D"].width = 15
ws2.column_dimensions["E"].width = 18

ws2.merge_cells("A1:E1")
ws2["A1"] = "RÉSUMÉ BUDGÉTAIRE"
ws2["A1"].font = Font(name="Calibri", bold=True, color=DARK_BLUE, size=16)
ws2["A1"].alignment = Alignment(horizontal="center")

row2 = 3
headers2 = ["N°", "Catégorie", "Montant (FCFA)", "%", "Montant (EUR)"]
for i, h in enumerate(headers2, 1):
    cell = ws2.cell(row=row2, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border

categories = [
    ("1", "Ressources Humaines", rh_total, cat_fills["RH"]),
    ("2", "Matériel (Hardware)", hw_total, cat_fills["HW"]),
    ("3", "Logiciels", sw_total, cat_fills["SW"]),
    ("4", "Frais de Fonctionnement", fonc_total, cat_fills["FONC"]),
]

for num, name, amount, fill in categories:
    row2 += 1
    ws2.cell(row=row2, column=1, value=num).alignment = center
    ws2.cell(row=row2, column=2, value=name).alignment = left_wrap
    ws2.cell(row=row2, column=3, value=amount).number_format = XAF_FORMAT
    ws2.cell(row=row2, column=3).alignment = right_align
    ws2.cell(row=row2, column=4, value=amount / total).number_format = PERCENT_FORMAT
    ws2.cell(row=row2, column=4).alignment = center
    ws2.cell(row=row2, column=5, value=round(amount / EUR_RATE)).number_format = '#,##0" EUR"'
    ws2.cell(row=row2, column=5).alignment = right_align
    for c in range(1, 6):
        ws2.cell(row=row2, column=c).font = normal_font
        ws2.cell(row=row2, column=c).border = thin_border
        ws2.cell(row=row2, column=c).fill = fill

# Total
row2 += 1
ws2.cell(row=row2, column=2, value="TOTAL").font = total_font
ws2.cell(row=row2, column=3, value=total).number_format = XAF_FORMAT
ws2.cell(row=row2, column=3).alignment = right_align
ws2.cell(row=row2, column=4, value=1).number_format = PERCENT_FORMAT
ws2.cell(row=row2, column=4).alignment = center
ws2.cell(row=row2, column=5, value=round(total / EUR_RATE)).number_format = '#,##0" EUR"'
ws2.cell(row=row2, column=5).alignment = right_align
style_range(ws2, row2, 1, 5, font=total_font, fill=total_fill, border=thin_border)

row2 += 1
ws2.cell(row=row2, column=2, value="Imprévus (10%)").font = bold_font
ws2.cell(row=row2, column=3, value=contingency).number_format = XAF_FORMAT
ws2.cell(row=row2, column=3).alignment = right_align
ws2.cell(row=row2, column=5, value=round(contingency / EUR_RATE)).number_format = '#,##0" EUR"'
ws2.cell(row=row2, column=5).alignment = right_align
style_range(ws2, row2, 1, 5, fill=cat_fills["gray"], border=thin_border)

row2 += 1
ws2.cell(row=row2, column=2, value="BUDGET TOTAL").font = total_font
ws2.cell(row=row2, column=3, value=grand_total).number_format = XAF_FORMAT
ws2.cell(row=row2, column=3).alignment = right_align
ws2.cell(row=row2, column=5, value=round(grand_total / EUR_RATE)).number_format = '#,##0" EUR"'
ws2.cell(row=row2, column=5).alignment = right_align
style_range(ws2, row2, 1, 5, font=total_font, fill=PatternFill(start_color="10B981", end_color="10B981", fill_type="solid"), border=thin_border)

# ─── Notes ───
row2 += 2
ws2.merge_cells(f"A{row2}:E{row2}")
ws2.cell(row=row2, column=1, value="NOTES ET ÉCONOMIES").font = subtitle_font

notes = [
    "✅ 8 logiciels open source utilisés = économie estimée > 2 000 000 FCFA",
    "✅ Certificat SSL gratuit via Let's Encrypt",
    "✅ RouterOS inclus avec le matériel MikroTik",
    "✅ Docker permet un déploiement sans coût de licence serveur",
    "⚠️ Les prix sont indicatifs et basés sur le marché camerounais (Jan. 2026)",
    "⚠️ La provision pour imprévus couvre les risques de dépassement",
]

for note in notes:
    row2 += 1
    ws2.merge_cells(f"A{row2}:E{row2}")
    ws2.cell(row=row2, column=1, value=note)
    ws2.cell(row=row2, column=1).font = Font(name="Calibri", size=11, color="64748B")

# Print settings
for sheet in [ws, ws2]:
    sheet.sheet_properties.pageSetUpPr = None
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = 9  # A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0

# Save
output = "/home/user/captive-portal/Budget_Portail_Captif_UCAC_ICAM.xlsx"
wb.save(output)
print(f"Budget saved to: {output}")
print(f"Total budget: {grand_total:,} FCFA ({round(grand_total / EUR_RATE):,} EUR)")
